"""Controlador para generación de reportes de proyección de cupos."""

from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from controllers.projection_controller import CourseProjection


class ReportController:
    """Genera reportes en Excel a partir de las proyecciones de cupos."""

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def generate_summary_report(
        self, projections: List[CourseProjection], plan_name: str, output_path: str
    ) -> str:
        """Genera un reporte resumen de cupos proyectados.

        Returns:
            Ruta del archivo generado.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen de Cupos"

        # Título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"Proyección de Cupos - {plan_name}"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Encabezados
        headers = ["Sigla", "Nombre del Curso", "Créditos", "Ciclo",
                    "Demanda Proyectada", "Ya Aprobaron", "Matriculados"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        # Datos
        for i, proj in enumerate(projections, 5):
            data = [
                proj.course.code,
                proj.course.name,
                proj.course.credits,
                proj.course.cycle,
                proj.projected_demand,
                proj.total_approved,
                proj.total_enrolled,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.font = self.BODY_FONT
                cell.border = self.BORDER
                if col >= 3:
                    cell.alignment = Alignment(horizontal="center")

        # Ajustar anchos
        widths = [12, 45, 10, 8, 18, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(output_path)
        return output_path

    def generate_detailed_report(
        self, projections: List[CourseProjection], plan_name: str, output_path: str
    ) -> str:
        """Genera un reporte detallado con listas de estudiantes por curso."""
        wb = openpyxl.Workbook()

        # Hoja resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        self._write_summary_sheet(ws_summary, projections, plan_name)

        # Hoja por cada curso con estudiantes elegibles
        for proj in projections:
            if proj.projected_demand == 0:
                continue
            sheet_name = proj.course.code[:31]
            ws = wb.create_sheet(title=sheet_name)
            self._write_course_detail_sheet(ws, proj)

        wb.save(output_path)
        return output_path

    def _write_summary_sheet(self, ws, projections, plan_name):
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Reporte Detallado - {plan_name}"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

        headers = ["Sigla", "Curso", "Ciclo", "Demanda", "Aprobados", "Matriculados"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER

        for i, proj in enumerate(projections, 4):
            values = [proj.course.code, proj.course.name, proj.course.cycle,
                      proj.projected_demand, proj.total_approved, proj.total_enrolled]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.font = self.BODY_FONT
                cell.border = self.BORDER

    def _write_course_detail_sheet(self, ws, proj: CourseProjection):
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"{proj.course.code} - {proj.course.name}"
        ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

        ws["A2"].value = f"Demanda proyectada: {proj.projected_demand}"
        ws["A2"].font = Font(name="Calibri", size=10, bold=True)

        headers = ["Carné", "Nombre", "Apellido 1", "Apellido 2"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER

        for i, student in enumerate(sorted(proj.eligible_students, key=lambda s: s.student_id), 5):
            values = [student.student_id, student.first_name,
                      student.last_name_1, student.last_name_2]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.font = self.BODY_FONT
                cell.border = self.BORDER

        for i, w in enumerate([15, 20, 20, 20], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
