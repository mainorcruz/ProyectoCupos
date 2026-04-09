"""Controlador para carga y procesamiento de archivos Excel de notas."""

from __future__ import annotations

from pathlib import Path
from typing import Tuple

import openpyxl

from models.student import (
    AcademicRecord,
    CourseRecord,
    Student,
    determine_status,
)


# Cursos de Arte y Repertorio: cualquiera de estas siglas equivale a ART100 o REP100
_ART_CODES = {
    "EG0124", "EG0125", "EG0126", "EG0127", "EG0128", "EG0129",
    "EG0130", "EG0131", "EG0132", "EG0133", "EG0134", "EG0135",
}
_REP_CODES = {
    "RP0013", "RP0017", "RP1109", "RP1111", "RP1112", "RP1113",
    "RP1117", "RP1202", "RP1237",
}


def _normalize_code(code: str) -> str:
    """Normaliza siglas de Cursos de Arte y Repertorio al código del plan."""
    bare = code.replace(" ", "")
    if bare in _ART_CODES:
        return "ART100"
    if bare in _REP_CODES:
        return "REP100"
    return code


# Mapeo de columnas esperadas
EXPECTED_COLUMNS = {
    "carne": "Carne",
    "apellido1": "Apellido 1",
    "apellido2": "Apellido 2",
    "nombre": "Nombre",
    "recinto": "Recinto",
    "sigla": "Sigla",
    "grupo": "Grupo",
    "nombre_curso": "Nombre del curso",
    "creditos": "Créditos",
    "nota": "Nota",
    "modalidad": "Modalidad",
    "pertenece_plan": "Pertenece a plan",
    "periodo": "Período",
    "anio": "Año",
}


class FileController:
    """Controla la carga de archivos Excel y construcción de expedientes."""

    def __init__(self):
        self.record = AcademicRecord()
        self._column_map: dict = {}

    def load_excel(self, file_path: str) -> Tuple[bool, str]:
        """Carga un archivo Excel y actualiza los expedientes estudiantiles.

        Args:
            file_path: Ruta al archivo .xlsx

        Returns:
            Tupla (éxito, mensaje)
        """
        path = Path(file_path)
        if not path.exists():
            return False, f"Archivo no encontrado: {file_path}"
        if path.suffix.lower() not in (".xlsx", ".xls"):
            return False, "El archivo debe ser .xlsx o .xls"

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active

            # Leer encabezados
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            self._column_map = self._map_columns(headers)

            if not self._column_map:
                return False, "No se encontraron las columnas esperadas en el archivo."

            missing = self._validate_required_columns()
            if missing:
                return False, f"Columnas faltantes: {', '.join(missing)}"

            rows_processed = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                values = [cell.value for cell in row]
                if not any(values):
                    continue
                self._process_row(values)
                rows_processed += 1

            wb.close()
            return True, (
                f"Archivo cargado exitosamente.\n"
                f"Filas procesadas: {rows_processed}\n"
                f"Estudiantes registrados: {self.record.student_count}"
            )

        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"

    def _map_columns(self, headers: list) -> dict:
        """Mapea las columnas del Excel a los campos internos."""
        col_map = {}
        header_lower = [h.lower().strip() for h in headers]

        aliases = {
            "carne": ["carne", "carné", "carn", "id"],
            "apellido1": ["apellido 1", "apellido1", "primer apellido"],
            "apellido2": ["apellido 2", "apellido2", "segundo apellido"],
            "nombre": ["nombre"],
            "recinto": ["recinto", "sede"],
            "sigla": ["sigla", "código", "codigo", "código curso"],
            "grupo": ["grupo", "grp"],
            "nombre_curso": ["nombre del curso", "nombre curso", "curso"],
            "creditos": ["créditos", "creditos", "créd"],
            "nota": ["nota", "calificación", "calificacion"],
            "modalidad": ["modalidad", "mod"],
            "pertenece_plan": ["pertenece a plan", "pertenece plan", "plan"],
            "periodo": ["período", "periodo", "per"],
            "anio": ["año", "anio", "year"],
        }

        for key, possible_names in aliases.items():
            for i, h in enumerate(header_lower):
                if h in possible_names:
                    col_map[key] = i
                    break

        return col_map

    def _validate_required_columns(self) -> list:
        required = ["carne", "sigla", "nota", "modalidad"]
        return [EXPECTED_COLUMNS.get(r, r) for r in required if r not in self._column_map]

    def _process_row(self, values: list) -> None:
        """Procesa una fila del Excel y actualiza el expediente."""
        get = lambda key: values[self._column_map[key]] if key in self._column_map else None

        student_id = str(get("carne") or "").strip()
        if not student_id:
            return

        first_name = str(get("nombre") or "").strip()
        last1 = str(get("apellido1") or "").strip()
        last2 = str(get("apellido2") or "").strip()
        campus = str(get("recinto") or "").strip()

        student = self.record.get_or_create_student(
            student_id, first_name, last1, last2, campus
        )

        code = _normalize_code(str(get("sigla") or "").strip().upper())
        if not code:
            return

        course_name = str(get("nombre_curso") or "").strip()
        credits_raw = get("creditos")
        credits = int(credits_raw) if credits_raw is not None else 0

        grade_raw = get("nota")
        grade = self._parse_grade(grade_raw)
        modality = str(get("modalidad") or "").strip()
        status = determine_status(grade, modality)

        period = str(get("periodo") or "").strip()
        year = str(get("anio") or "").strip()
        group = str(get("grupo") or "").strip()
        belongs_plan = str(get("pertenece_plan") or "").strip()

        record = CourseRecord(
            code=code,
            name=course_name,
            credits=credits,
            grade=grade,
            status=status,
            modality=modality,
            period=period,
            year=year,
            group=group,
            belongs_to_plan=belongs_plan,
        )
        student.add_record(record)

    @staticmethod
    def _parse_grade(value) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip().replace(",", "."))
        except (ValueError, TypeError):
            return None

    def reset(self) -> None:
        """Reinicia los expedientes."""
        self.record = AcademicRecord()
        self._column_map = {}
