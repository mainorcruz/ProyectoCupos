"""Vista principal de la aplicación de proyección de cupos."""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Optional

from models.plan_data import PLAN_01, PLAN_07
from models.study_plan import StudyPlan
from controllers.file_controller import FileController
from controllers.projection_controller import ProjectionController, CourseProjection
from controllers.report_controller import ReportController


class MainView:
    """Ventana principal de la aplicación."""

    WINDOW_TITLE = "Proyección de Cupos - Bachillerato Informática Empresarial (UCR)"
    WINDOW_SIZE = "1100x720"
    BG_COLOR = "#F0F4F8"
    PRIMARY = "#1F4E79"
    ACCENT = "#2E86C1"
    SUCCESS = "#27AE60"

    def __init__(self):
        self.root = tk.Tk()
        self.root.title(self.WINDOW_TITLE)
        self.root.geometry(self.WINDOW_SIZE)
        self.root.configure(bg=self.BG_COLOR)
        self.root.minsize(900, 600)

        self.file_controller = FileController()
        self.projection_controller: Optional[ProjectionController] = None
        self.report_controller = ReportController()
        self.current_plan: Optional[StudyPlan] = None

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"),
                        foreground=self.PRIMARY, background=self.BG_COLOR)
        style.configure("Subtitle.TLabel", font=("Segoe UI", 10),
                        foreground="#555", background=self.BG_COLOR)
        style.configure("Section.TLabelframe.Label", font=("Segoe UI", 10, "bold"),
                        foreground=self.PRIMARY)
        style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=24)
        style.configure("Status.TLabel", font=("Segoe UI", 9),
                        background="#E8EEF2", foreground="#333")

    def _build_ui(self):
        # Header
        header = tk.Frame(self.root, bg=self.PRIMARY, height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="📊 Proyección de Cupos", font=("Segoe UI", 18, "bold"),
                 fg="white", bg=self.PRIMARY).pack(side=tk.LEFT, padx=20, pady=10)
        tk.Label(header, text="Bachillerato en Informática Empresarial — UCR",
                 font=("Segoe UI", 10), fg="#B0C4DE", bg=self.PRIMARY).pack(side=tk.LEFT)

        # Main content
        main_frame = tk.Frame(self.root, bg=self.BG_COLOR)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        # Left panel - controls
        left = tk.Frame(main_frame, bg=self.BG_COLOR, width=340)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left.pack_propagate(False)

        self._build_file_section(left)
        self._build_plan_section(left)
        self._build_filter_section(left)
        self._build_actions_section(left)

        # Right panel - results
        right = tk.Frame(main_frame, bg=self.BG_COLOR)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_results_section(right)

        # Status bar
        self.status_var = tk.StringVar(value="Listo. Cargue un archivo Excel para comenzar.")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, style="Status.TLabel",
                               relief=tk.SUNKEN, anchor=tk.W, padding=(10, 4))
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _build_file_section(self, parent):
        frame = ttk.LabelFrame(parent, text="  1. Cargar Archivo de Notas  ",
                                style="Section.TLabelframe", padding=10)
        frame.pack(fill=tk.X, pady=(0, 10))

        self.file_label = ttk.Label(frame, text="Ningún archivo seleccionado",
                                     wraplength=280, foreground="#888")
        self.file_label.pack(anchor=tk.W, pady=(0, 5))

        btn_frame = tk.Frame(frame, bg=self.BG_COLOR)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="📁 Seleccionar Excel",
                   command=self._on_load_file).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(btn_frame, text="🔄", width=3,
                   command=self._on_reset).pack(side=tk.RIGHT, padx=(5, 0))

        self.load_info = ttk.Label(frame, text="", foreground=self.SUCCESS, wraplength=280)
        self.load_info.pack(anchor=tk.W, pady=(5, 0))

    def _build_plan_section(self, parent):
        frame = ttk.LabelFrame(parent, text="  2. Plan de Estudios  ",
                                style="Section.TLabelframe", padding=10)
        frame.pack(fill=tk.X, pady=(0, 10))

        self.plan_var = tk.StringVar(value="01")
        plans = [("Plan 01 — Plan de 1997", "01"),
                 ("Plan 07 — Rediseño", "07")]
        for text, val in plans:
            ttk.Radiobutton(frame, text=text, variable=self.plan_var,
                           value=val, command=self._on_plan_changed).pack(anchor=tk.W, pady=2)

    def _build_filter_section(self, parent):
        frame = ttk.LabelFrame(parent, text="  3. Seleccionar Cursos  ",
                                style="Section.TLabelframe", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Cycle filter
        cycle_frame = tk.Frame(frame, bg=self.BG_COLOR)
        cycle_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(cycle_frame, text="Ciclo:").pack(side=tk.LEFT)
        self.cycle_var = tk.StringVar(value="Todos")
        self.cycle_combo = ttk.Combobox(cycle_frame, textvariable=self.cycle_var,
                                         state="readonly", width=10)
        self.cycle_combo.pack(side=tk.LEFT, padx=5)
        self.cycle_combo.bind("<<ComboboxSelected>>", self._on_cycle_changed)

        # Select all / none
        sel_frame = tk.Frame(frame, bg=self.BG_COLOR)
        sel_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(sel_frame, text="Todos", command=self._select_all, width=8).pack(side=tk.LEFT)
        ttk.Button(sel_frame, text="Ninguno", command=self._select_none, width=8).pack(side=tk.LEFT, padx=5)

        # Course listbox with checkboxes
        list_frame = tk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.course_listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE,
                                          font=("Consolas", 9),
                                          yscrollcommand=scrollbar.set,
                                          activestyle="none")
        self.course_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.course_listbox.yview)

        self._course_codes: list = []
        self._update_plan()

    def _build_actions_section(self, parent):
        frame = ttk.LabelFrame(parent, text="  4. Generar Reporte  ",
                                style="Section.TLabelframe", padding=10)
        frame.pack(fill=tk.X)

        ttk.Button(frame, text="▶ Proyectar Cupos",
                   command=self._on_project, style="Accent.TButton").pack(fill=tk.X, pady=(0, 5))

        export_frame = tk.Frame(frame, bg=self.BG_COLOR)
        export_frame.pack(fill=tk.X)
        ttk.Button(export_frame, text="💾 Resumen Excel",
                   command=self._on_export_summary).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(export_frame, text="📋 Detallado Excel",
                   command=self._on_export_detailed).pack(side=tk.RIGHT, expand=True, fill=tk.X, padx=(5, 0))

    def _build_results_section(self, parent):
        frame = ttk.LabelFrame(parent, text="  Resultados de Proyección  ",
                                style="Section.TLabelframe", padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        columns = ("code", "name", "credits", "cycle", "demand", "approved", "enrolled")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse")

        headings = {
            "code": ("Sigla", 80),
            "name": ("Nombre del Curso  🔍", 260),
            "credits": ("Créd.", 50),
            "cycle": ("Ciclo", 50),
            "demand": ("Demanda", 80),
            "approved": ("Aprobaron", 80),
            "enrolled": ("Matriculados", 90),
        }
        for col, (text, width) in headings.items():
            self.tree.heading(col, text=text,
                              command=lambda c=col: self._sort_treeview(c))
            self.tree.column(col, width=width, anchor=tk.CENTER if col != "name" else tk.W)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self.tree.bind("<Motion>", self._on_tree_motion)

        # Hint
        ttk.Label(parent, text="Clic en el nombre del curso para ver la lista de estudiantes pendientes.",
                  font=("Segoe UI", 8), foreground="#888",
                  background=self.BG_COLOR).pack(anchor=tk.W, pady=(3, 0))

        # Summary bar
        self.summary_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.summary_var, font=("Segoe UI", 10, "bold"),
                  foreground=self.PRIMARY, background=self.BG_COLOR).pack(anchor=tk.W, pady=(2, 0))

        self._last_projections: list = []
        self._sort_reverse = {}

    # ── Event handlers ───────────────────────────────────────────────

    def _on_load_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo de notas",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not path:
            return

        self.file_controller.reset()
        success, msg = self.file_controller.load_excel(path)

        if success:
            short_name = path.split("/")[-1].split("\\")[-1]
            self.file_label.config(text=short_name, foreground=self.PRIMARY)
            self.load_info.config(text=msg, foreground=self.SUCCESS)
            self.status_var.set(f"Archivo cargado: {short_name}")
        else:
            self.file_label.config(text="Error", foreground="red")
            self.load_info.config(text=msg, foreground="red")
            messagebox.showerror("Error", msg)

    def _on_reset(self):
        self.file_controller.reset()
        self.file_label.config(text="Ningún archivo seleccionado", foreground="#888")
        self.load_info.config(text="")
        self._clear_results()
        self.status_var.set("Datos reiniciados.")

    def _on_plan_changed(self):
        self._update_plan()
        self._clear_results()

    def _on_cycle_changed(self, event=None):
        self._populate_course_list()

    def _on_project(self):
        if self.file_controller.record.student_count == 0:
            messagebox.showwarning("Sin datos", "Primero cargue un archivo Excel de notas.")
            return

        selected_indices = self.course_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Sin selección", "Seleccione al menos un curso.")
            return

        selected_codes = [self._course_codes[i] for i in selected_indices]
        self.projection_controller = ProjectionController(
            self.current_plan, self.file_controller.record
        )

        projections = self.projection_controller.project_selected_courses(selected_codes)
        self._display_projections(projections)

    def _on_export_summary(self):
        self._export_report(detailed=False)

    def _on_export_detailed(self):
        self._export_report(detailed=True)

    def _export_report(self, detailed: bool):
        if not self._last_projections:
            messagebox.showwarning("Sin datos", "Primero genere una proyección.")
            return

        suffix = "detallado" if detailed else "resumen"
        path = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".xlsx",
            initialfile=f"reporte_cupos_{suffix}_{self.current_plan.plan_id}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            plan_name = f"Plan {self.current_plan.plan_id} — {self.current_plan.name}"
            if detailed:
                self.report_controller.generate_detailed_report(
                    self._last_projections, plan_name, path
                )
            else:
                self.report_controller.generate_summary_report(
                    self._last_projections, plan_name, path
                )
            messagebox.showinfo("Éxito", f"Reporte guardado en:\n{path}")
            self.status_var.set(f"Reporte exportado: {path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")

    # ── UI helpers ───────────────────────────────────────────────────

    def _update_plan(self):
        plan_id = self.plan_var.get()
        self.current_plan = PLAN_01 if plan_id == "01" else PLAN_07
        cycles = ["Todos"] + [str(c) for c in self.current_plan.get_all_cycles()]
        self.cycle_combo["values"] = cycles
        self.cycle_var.set("Todos")
        self._populate_course_list()

    def _populate_course_list(self):
        self.course_listbox.delete(0, tk.END)
        self._course_codes.clear()

        cycle_filter = self.cycle_var.get()
        courses = list(self.current_plan.courses.values())
        courses.sort(key=lambda c: (c.cycle, c.code))

        if cycle_filter != "Todos":
            cycle_num = int(cycle_filter)
            courses = [c for c in courses if c.cycle == cycle_num]

        for course in courses:
            label = f"C{course.cycle:>2} │ {course.code:<10} │ {course.name}"
            self.course_listbox.insert(tk.END, label)
            self._course_codes.append(course.code)

    def _select_all(self):
        self.course_listbox.select_set(0, tk.END)

    def _select_none(self):
        self.course_listbox.select_clear(0, tk.END)

    def _display_projections(self, projections: list):
        self._last_projections = projections
        self.tree.delete(*self.tree.get_children())

        total_demand = 0
        for proj in projections:
            self.tree.insert("", tk.END, values=(
                proj.course.code,
                proj.course.name,
                proj.course.credits,
                proj.course.cycle,
                proj.projected_demand,
                proj.total_approved,
                proj.total_enrolled,
            ))
            total_demand += proj.projected_demand

        self.summary_var.set(
            f"Cursos: {len(projections)}  |  "
            f"Demanda total proyectada: {total_demand}  |  "
            f"Estudiantes en expediente: {self.file_controller.record.student_count}"
        )
        self.status_var.set(f"Proyección completada — {len(projections)} cursos analizados.")

    def _clear_results(self):
        self.tree.delete(*self.tree.get_children())
        self.summary_var.set("")
        self._last_projections.clear() if self._last_projections else None

    def _sort_treeview(self, col):
        reverse = self._sort_reverse.get(col, False)
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]

        try:
            items.sort(key=lambda t: float(t[0]), reverse=reverse)
        except ValueError:
            items.sort(key=lambda t: t[0].lower(), reverse=reverse)

        for idx, (_, k) in enumerate(items):
            self.tree.move(k, "", idx)

        self._sort_reverse[col] = not reverse

    def _on_tree_motion(self, event):
        """Cambia el cursor a 'mano' cuando el mouse está sobre la columna de nombre."""
        col = self.tree.identify_column(event.x)
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell" and col == "#2":
            self.tree.configure(cursor="hand2")
        else:
            self.tree.configure(cursor="")

    def _on_tree_click(self, event):
        """Abre ventana de detalle al hacer clic en el nombre del curso."""
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#2":
            return
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        course_code = self.tree.item(row_id, "values")[0]
        proj = next((p for p in self._last_projections if p.course.code == course_code), None)
        if proj:
            self._show_student_detail(proj)

    def _show_student_detail(self, proj):
        """Muestra ventana con lista de estudiantes pendientes del curso y opción de exportar."""
        students = sorted(proj.eligible_students, key=lambda s: s.student_id)

        win = tk.Toplevel(self.root)
        win.title(f"Estudiantes pendientes — {proj.course.code}")
        win.geometry("620x480")
        win.configure(bg=self.BG_COLOR)
        win.transient(self.root)
        win.grab_set()

        # Header
        header = tk.Frame(win, bg=self.PRIMARY, height=55)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text=f"{proj.course.code}  —  {proj.course.name}",
                 font=("Segoe UI", 13, "bold"), fg="white", bg=self.PRIMARY
                 ).pack(side=tk.LEFT, padx=15, pady=10)

        # Barra resumen
        summary_frame = tk.Frame(win, bg="#E8EEF2")
        summary_frame.pack(fill=tk.X)
        tk.Label(summary_frame,
                 text=f"  Estudiantes pendientes: {len(students)}",
                 font=("Segoe UI", 9, "bold"), bg="#E8EEF2", fg=self.PRIMARY, pady=6
                 ).pack(side=tk.LEFT, padx=10)

        # Tabla de estudiantes
        table_frame = tk.Frame(win)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(8, 4))

        cols = ("id", "nombre", "apellido1", "apellido2")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")

        for col, text, width, anchor in [
            ("id",        "Carné",      100, tk.CENTER),
            ("nombre",    "Nombre",     150, tk.W),
            ("apellido1", "Apellido 1", 150, tk.W),
            ("apellido2", "Apellido 2", 150, tk.W),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor=anchor)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        for student in students:
            tree.insert("", tk.END, values=(
                student.student_id, student.first_name,
                student.last_name_1, student.last_name_2
            ))

        # Botón exportar
        btn_frame = tk.Frame(win, bg=self.BG_COLOR)
        btn_frame.pack(fill=tk.X, padx=10, pady=(4, 10))
        ttk.Button(
            btn_frame, text="💾 Exportar a Excel",
            command=lambda: self._export_student_list(proj, students)
        ).pack(side=tk.RIGHT)

    def _export_student_list(self, proj, students):
        """Exporta la lista de estudiantes pendientes a Excel."""
        path = filedialog.asksaveasfilename(
            title="Guardar lista de estudiantes",
            defaultextension=".xlsx",
            initialfile=f"pendientes_{proj.course.code}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pendientes"

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            body_font   = Font(name="Calibri", size=11)
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"),
            )

            # Título
            ws.merge_cells("A1:D1")
            ws["A1"].value = f"Estudiantes pendientes — {proj.course.code}: {proj.course.name}"
            ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
            ws["A1"].alignment = Alignment(horizontal="left")

            ws["A2"].value = f"Total: {len(students)} estudiantes"
            ws["A2"].font = Font(name="Calibri", size=10, italic=True)

            # Encabezados
            for col, text in enumerate(["Carné", "Nombre", "Apellido 1", "Apellido 2"], 1):
                cell = ws.cell(row=4, column=col, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Datos
            for i, student in enumerate(students, 5):
                for col, val in enumerate([
                    student.student_id, student.first_name,
                    student.last_name_1, student.last_name_2
                ], 1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.font = body_font
                    cell.border = border

            for i, w in enumerate([15, 20, 20, 20], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            wb.save(path)
            messagebox.showinfo("Éxito", f"Lista exportada:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")

    def run(self):
        self.root.mainloop()
